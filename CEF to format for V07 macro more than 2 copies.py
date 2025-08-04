# -*- coding: utf-8 -*-
"""
Created on Wed Apr  9 14:34:19 2025

@author: ymohdzaifullizan
"""

import os
import openpyxl

def modify_sheets(file_path, output_file):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)  # Load workbook with 'data_only' to get calculated values
    # Create a new workbook for the output
    new_workbook = openpyxl.Workbook()
    # Remove the default sheet created by openpyxl
    new_workbook.remove(new_workbook.active)
    
    # Iterate over all sheets in the workbook starting with "CEF"
    sheets_to_modify = [sheet for sheet in workbook.sheetnames if sheet.startswith("CEF")]
    
    for sheet_name in sheets_to_modify:
        original_sheet = workbook[sheet_name]
        # Check if cell B6 is empty
        if not original_sheet["B6"].value:
            print(f"Skipping sheet '{sheet_name}' because cell B6 is empty.")
            continue  # Skip this sheet
        
        # Delete row 25 from the original sheet
        print(f"Deleting row 25 in the sheet: {sheet_name}")
        original_sheet.delete_rows(25)
        
        # Define rows to check in column B
        rows_to_check = [6, 7, 8, 9, 10]
        
        for row_number in rows_to_check:
            # Check if the corresponding cell in column B is non-empty
            cell_value = original_sheet[f"B{row_number}"].value
            if not cell_value:
                print(f"Skipping creation of copy for row {row_number} because B{row_number} is empty.")
                continue
            
            # Create a copy for this row
            copy_sheet = new_workbook.create_sheet(title=f"{sheet_name}_Copy{row_number - 5}")  # Copy names: Copy1, Copy2, ...
            copy_sheet_values_only(original_sheet, copy_sheet)  # Copy values only from original sheet
            
            # Delete all rows except the current one being checked
            rows_to_delete = [r for r in rows_to_check if r != row_number]  # Keep only the current row
            delete_specific_rows(copy_sheet, rows_to_delete)
            
            # Clear contents of rows 15-19 #delete this section because the raw part numbers are not referenced in the Paint file, we are only interested in the painted parts list.
            clear_row_contents(copy_sheet, [15, 16, 17, 18, 19])
    
    # Save the new workbook to the provided output path
    new_workbook.save(output_file)
    print(f"Modified Excel file saved as: {output_file}")


def copy_sheet_values_only(original_sheet, copy_sheet):
    """
    Copies the content of the original sheet into a new sheet, replacing formulas with calculated values.
    Arguments:
    original_sheet -- openpyxl worksheet object (source)
    copy_sheet -- openpyxl worksheet object (destination)
    """
    for row in original_sheet.iter_rows():  # Iterate through rows with formulas and values
        new_row = []
        for cell in row:
            new_row.append(cell.value)  # Extract the calculated value
        copy_sheet.append(new_row)  # Append the values into the destination sheet


def delete_specific_rows(sheet, rows_to_delete):
    """
    Deletes specific rows from the sheet while ensuring row indices to delete don't affect subsequent operations.
    Arguments:
    sheet -- openpyxl worksheet object
    rows_to_delete -- List of row indices to delete completely
    """
    rows_to_delete = sorted(rows_to_delete, reverse=True)  # Delete from bottom to top
    for row in rows_to_delete:
        sheet.delete_rows(row)


def clear_row_contents(sheet, rows_to_clear):
    """
    Clears the contents of specific rows without deleting the rows themselves.
    Arguments:
    sheet -- openpyxl worksheet object
    rows_to_clear -- List of row indices whose contents should be cleared
    """
    for row in rows_to_clear:
        for cell in sheet[row]:  # Iterate through all cells in the row
            cell.value = None  # Set each cell's value to None (clear contents)


def process_folder(input_folder, output_folder):
    """
    Processes all Excel files in a folder and saves the modified copies to a new folder.
    Arguments:
    input_folder -- The folder containing the original Excel files
    output_folder -- The folder to save the modified Excel files
    """
    # Create the output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Loop through all Excel files in the input folder
    for file_name in os.listdir(input_folder):
        if file_name.endswith(".xlsx"):  # Skip non-Excel files
            input_file_path = os.path.join(input_folder, file_name)
            output_file_path = os.path.join(output_folder, f"Modified_{file_name}")
            
            print(f"Processing file: {input_file_path}")
            modify_sheets(input_file_path, output_file_path)

# --- Run the folder processing function ---
input_folder = r"C:\Users\ymohdzaifullizan\OneDrive - Dyson\Year 2 rotation - E&O\Paint\8. August 25\11.1 Luxshare CEF needs mods"
output_folder = r"C:\Users\ymohdzaifullizan\OneDrive - Dyson\Year 2 rotation - E&O\Paint\8. August 25\11.2 Luxshare CEF modified"

process_folder(input_folder, output_folder)